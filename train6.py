from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from time import sleep
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import pandas as ps
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
  

if __name__=='__main__':
    data= ps.read_excel("/Users/brahim/Downloads/Python-Selenium-Excel-Test-master/Classeur1.xlsx")
    print("data count is \n", data['login'].values)
    print("data requested is \n", data['password'].values)
    donner=data['id'].count()
    print(data)
    expected_title="Welcome - LambdaTest"
    t=[]
    
    for i in range(donner):
        service=Service(r"C:\Users\HP\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
        options=webdriver.FirefoxOptions()
        browser=webdriver.Firefox()
        browser.get("https://accounts.lambdatest.com/login")
        try:
            li1=WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, 'email')))
            print(data['login'].values[i])
            print(data['password'].values[i])
            if(ps.isna(data['login'].values[i]) ):
               li1.send_keys("")
               #print("hadi2")
            else:
                li1.send_keys(data['login'].values[i])
                #print("hadi1")
                
            li1.click()

            li2=WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, "password")))
            if(ps.isna(data['password'].values[i])):
                #print("hadi2")
                li2.send_keys("")
            else:
                #print("hadi1")
                li2.send_keys(data['password'].values[i])
                
            li2.click()

            li3=WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, 'login-button')))
            li3.click()
            
            sleep(5)
            actual_title= browser.title
            print(actual_title," vs ",expected_title)
            if( actual_title == expected_title):
                data.loc[i, 'results']='OK'
                
                data.style.highlight_null('red')
                print(data)
                t.append("OK")
            else:
                data.loc[i, 'results']='KO'
                data.style.highlight_null( 'bleu')
                t.append("KO")
           
            data.to_excel("/Users/brahim/Downloads/Python-Selenium-Excel-Test-master/Classeur7.xlsx", index=False)
            ### -------------colorisation----------------- ####
            # Load the workbook and select the active worksheet
            wb = load_workbook("/Users/brahim/Downloads/Python-Selenium-Excel-Test-master/Classeur7.xlsx")
            ws = wb.active

            # Define the fill colors
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

            # Iterate over the cells in the DataFrame
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):  # Skip header row
                for cell in row:
                    if cell.value == "KO":
                        cell.fill = red_fill
                    elif cell.value == "OK":
                        cell.fill = green_fill

            # Save the changes to the new Excel file
            wb.save("/Users/brahim/Downloads/Python-Selenium-Excel-Test-master/Classeur7.xlsx")
            ### ------------------------------------------------------###
            browser.quit()
        except TimeoutException:
            print("Element Not Found")
        sleep(5)
        browser.quit()
        #print(t)
    data['results']=t
    #data.to_excel("C:/Users/HP/Downloads/Classeur12.xlsx", index=False)
    